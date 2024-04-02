# Import des librairies
import pandas as pd
import numpy as np
import sys
import os
import pickle
from keras.models import load_model
import streamlit as st
st.set_page_config(layout="wide")

import matplotlib.pyplot as plt
from PIL import Image
import requests
from io import BytesIO
from bs4 import BeautifulSoup

# Pour pouvoir réaliser des imports
sys.path.append(os.path.abspath(r"..\data\utils"))
sys.path.append(os.path.abspath(r"..\src\utils"))

from img_to_excel_func import imgtoexcel
from helper_functions import *

# Titre de la page
st.title("Détection de défauts")

# Le dossier source des images

image_dir = st.sidebar.text_input(r"répertoire des images (exemple : ..\data\train)")
# Exemple : ..\data\train

if image_dir.startswith('http'):
    url = image_dir#"http://192.168.0.20:9090/DCIM/Camera" #"http://192.168.0.XX:8000"  # Assuming this is the correct URL
    response = requests.get(url)
    
    if response.status_code == 200:
        # If the request was successful, you can access the contents of the directory
        directory_contents = response.text
        
    else:
        print("Failed to retrieve directory contents. Status code:", response.status_code)

    soup = BeautifulSoup(directory_contents, 'html.parser')

    # Extract URLs of the images
    image_urls = [url.get('href') for url in soup.find_all('a')]
    print('img_url : ', image_urls)

    # Créer un dossier Images dans le répertoire courant et y stocker les images
    os.makedirs("Images",exist_ok=True)
    for i,img in enumerate(image_urls[5:]):
        try: 
            name = os.path.basename(img)
                
            chosen_image_url = img #image_urls[2]  # Change index to select a different image
            
            # URL de l'image que vous souhaitez afficher
            chosen_image_url = str(url+'/'+chosen_image_url) #"http://192.168.0.29:8000/eurocast_20220311_090618.jpg"
            
            # Ouvrir l'image depuis l'URL
            response = requests.get(chosen_image_url)
            img = Image.open(BytesIO(response.content))
            
            # Convertir l'image en un tableau NumPy
            img_np = np.array(img)
    
            # Sauvegarder l'image au format JPG dans un dossier Images du répertoire courant
            img_jpg = Image.fromarray(img_np)
            print('l image est : ',img)
            img_jpg.save(f"Images/{name}")
                    
        except:
            print(img,' is not a jpg')
            continue
        # donne à image_dir le nouveau nom Images (miroir du http)
        image_dir = "Images"

# Le document excel de sortie

result_doc = st.sidebar.text_input(r"document de sortie (exemple : ..\data\train_excel.xlsx)")

# Le bouton de prédiction

if st.button("Analyse des images"):
    model = load_model(r"..\models\trained_model.h5")
    data = imgtoexcel(image_dir, result_doc)   
    
    preprocessing_entry = preprocess_input
    test_flow = define_flow(data, preprocessing_entry)
    y_pred_cat, y_pred = performance_test(data, model, test_flow)

    # Ajout de la colonne de prédiction au doc excel
    from openpyxl import load_workbook

    # Charger le document Excel existant
    workbook = load_workbook(result_doc)
    
    # Sélectionner la feuille de calcul sur laquelle vous souhaitez ajouter la colonne
    sheet = workbook.active  # Ou utilisez workbook['Nom_de_la_feuille']
    
    # Ajouter une colonne à la feuille de calcul
    new_column = ['Pred']+list(y_pred_cat) # Remplacez ceci par les valeurs que vous souhaitez ajouter
    sheet.insert_cols(idx=10, amount=1)  # Insérer une colonne à l'indice 9 (colonne J)
    
    # Écrire les valeurs dans la nouvelle colonne
    for i, value in enumerate(new_column, start=1):
        sheet.cell(row=i, column=10, value=value)  # Colonne C, changez le numéro de colonne selon votre besoin
    
    # Enregistrer les modifications
    workbook.save(result_doc)

    #Ajout d'une colonne pred au dataframe
    df = data.copy()
    df['pred'] = y_pred_cat
    #df.to_excel(result_doc, index=False)
        
    # Affichage des pièces non conformes 
    for img_path in df.loc[df['pred']=='nc','Image_Path'].tolist():
        img = cv2.imread(img_path)
        img = cv2.resize(img, (256,256))
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        st.text(img_path)
        st.image(img)
